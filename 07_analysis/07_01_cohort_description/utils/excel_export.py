import os
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .table_model import (
    KIND_COHORT,
    KIND_CONT,
    KIND_SECTION,
    KIND_SUB,
    TableModel,
)

FONT_NAME = "Arial"

_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
_HEADER_FONT = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
_SECTION_FONT = Font(name=FONT_NAME, bold=True, size=10)
_LABEL_FONT = Font(name=FONT_NAME, bold=True, size=10)
_BODY_FONT = Font(name=FONT_NAME, size=10)

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
_COHORT_FILL = PatternFill("solid", fgColor="F2F2F2")

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WRAP_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_TOP_LEFT = Alignment(horizontal="left", vertical="top")
_TOP_LEFT_INDENT = Alignment(horizontal="left", vertical="top", indent=2)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, model: TableModel) -> None:
    widths = [0] * model.ncols
    grid = [model.header] + [r.cells for r in model.rows]
    for row in grid:
        for j, val in enumerate(row):
            for line in str(val).split("\n"):
                widths[j] = max(widths[j], len(line))
    # first column holds the characteristic names -> a bit wider
    for j, w in enumerate(widths):
        col = get_column_letter(j + 1)
        ws.column_dimensions[col].width = min(max(w + 3, 12), 55)


def _write_sheet(wb: Workbook, model: TableModel) -> None:
    ws = wb.create_sheet(title=model.sheet_name[:31])
    ncols = model.ncols

    # --- title ----------------------------------------------------------- #
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=model.title)
    tcell.font = _TITLE_FONT
    tcell.alignment = _TOP_LEFT

    # --- column header --------------------------------------------------- #
    hr = 2
    for j, txt in enumerate(model.header, start=1):
        c = ws.cell(row=hr, column=j, value=txt)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER
    ws.row_dimensions[hr].height = 34

    # --- body ------------------------------------------------------------ #
    r = hr + 1
    for row in model.rows:
        for j, val in enumerate(row.cells, start=1):
            c = ws.cell(row=r, column=j, value=val if val != "" else None)
            c.border = _BORDER
            if row.kind == KIND_SECTION:
                c.font = _SECTION_FONT
                c.fill = _SECTION_FILL
                c.alignment = _TOP_LEFT
            elif row.kind == KIND_COHORT:
                c.font = _BODY_FONT
                c.fill = _COHORT_FILL
                c.alignment = _WRAP_TOP_LEFT
            elif row.kind == KIND_SUB:
                c.font = _BODY_FONT
                c.alignment = _TOP_LEFT_INDENT if j == 1 else _TOP_LEFT
            elif row.kind == KIND_CONT:
                c.font = _BODY_FONT
                c.alignment = _TOP_LEFT
            else:  # KIND_DATA
                c.font = _LABEL_FONT if j == 1 else _BODY_FONT
                c.alignment = _TOP_LEFT
        if row.kind == KIND_COHORT:
            ws.row_dimensions[r].height = 30
        r += 1

    _autosize(ws, model)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"


def write_workbook(
    tables: Sequence[TableModel],
    path: str,
    methods_model: TableModel | None = None,
) -> str:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    for model in tables:
        _write_sheet(wb, model)
    if methods_model is not None:
        _write_sheet(wb, methods_model)

    wb.save(path)
    return path
