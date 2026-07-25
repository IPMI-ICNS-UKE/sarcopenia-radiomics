from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from config import CASE_ORDER, RCFG, TableConfig


# Column specification
_COLUMN_SPEC: List[tuple] = [
    # Group 1 – identifier
    ("patient_id", "Patient", 1, None),
    ("case_type", "Case type", 1, None),
    # Group 2 – demographics
    ("sex", "Sex (1 = male)", 2, "int"),
    ("age", "Age, years", 2, "int"),
    ("bmi", "BMI", 2, "float:2"),
    ("height", "Height, m", 2, "float:2"),
    # Group 3 – functional tests
    ("hand_grip_cont", "Hand grip strength, kg", 3, "float:2"),
    ("chair_rise_cont", "Chair rise time, s", 3, "float:2"),
    ("gait_speed_cont", "Gait speed, m/s", 3, "float:2"),
    # Group 4 – model output
    ("musclefat_score", "Muscle fat score", 4, "float:2"),
    ("ct_score", "CT score", 4, "float:2"),
    ("combined_score", "Combined score", 4, "float:2"),
    ("pred_proba", "Predicted probability", 4, "float:2"),
    ("model_threshold", "Model threshold", 4, "float:2"),
]


def _fmt(value: Any, spec: Optional[str]) -> Any:
    """Format a single value according to its column spec."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    if spec is None:
        return value
    if spec == "int":
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    if spec.startswith("float:"):
        decimals = int(spec.split(":")[1])
        try:
            return round(float(value), decimals)
        except (TypeError, ValueError):
            return value
    return value


def build_case_table(
    case_records: Dict[str, Dict[str, Any]],
) -> pd.DataFrame:
    rows = []
    for case_type in CASE_ORDER:
        rec = case_records.get(case_type, {})
        row: Dict[str, Any] = {}
        for key, display, _grp, fmt in _COLUMN_SPEC:
            if key == "case_type":
                row[display] = case_type
            else:
                row[display] = _fmt(rec.get(key), fmt)
        rows.append(row)
    return pd.DataFrame(rows)


def save_excel_table(
    df: pd.DataFrame,
    path: Path,
    table_cfg: TableConfig = RCFG.table,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    if not _HAS_OPENPYXL:
        csv_path = path.with_suffix(".csv")
        df.to_csv(csv_path, index=False)
        print(f"  [openpyxl not installed] Saved as CSV: {csv_path}")
        return

    # Map display name → colour group
    col_to_group: Dict[str, int] = {display: grp for _, display, grp, _ in _COLUMN_SPEC}
    group_fills = {
        1: PatternFill("solid", fgColor=table_cfg.color_group1),
        2: PatternFill("solid", fgColor=table_cfg.color_group2),
        3: PatternFill("solid", fgColor=table_cfg.color_group3),
        4: PatternFill("solid", fgColor=table_cfg.color_group4),
    }
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"

    # ---- Header row --------------------------------------------------------
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center
        grp = col_to_group.get(col_name, 1)
        cell.fill = group_fills[grp]

    # ---- Data rows ---------------------------------------------------------
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, (col_name, value) in enumerate(zip(df.columns, row_data), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center
            grp = col_to_group.get(col_name, 1)
            cell.fill = group_fills[grp]

    # ---- Auto-width --------------------------------------------------------
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(path)
    print(f"  Saved table: {path}")
