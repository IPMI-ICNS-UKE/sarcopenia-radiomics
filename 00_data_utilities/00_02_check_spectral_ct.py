from pathlib import Path
import re
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


IMAGES_ROOT = Path("/home/gkolokolnikov/PhD_project/vault_data/SarcopeniaDataLiver/imagesAll")
OUTPUT_XLSX = Path("/home/gkolokolnikov/PhD_project/vault_data/SarcopeniaDataLiver/summary/summary_images.xlsx")
EXPECTED_PHASES = ["arteriell", "nativ", "spaet", "venous", "non_def"]


def is_nii_gz(path: Path) -> bool:
    return path.is_file() and path.name.lower().endswith(".nii.gz")


def patient_sort_key(path: Path) -> Tuple[int, str]:
    """
    Sort Pat folders numerically where possible:
    Pat1, Pat2, Pat10, ...
    """
    match = re.match(r"^Pat(\d+)$", path.name)
    if match:
        return (int(match.group(1)), path.name)
    return (10**12, path.name)


def count_nii_gz_files(folder: Path) -> int:
    if not folder.exists() or not folder.is_dir():
        return 0
    return sum(1 for p in folder.iterdir() if is_nii_gz(p))


def collect_summary(images_root: Path):
    overview_rows: List[List] = []
    long_rows: List[List] = []
    unexpected_rows: List[List] = []

    if not images_root.exists():
        raise FileNotFoundError(f"Images root does not exist: {images_root}")

    patient_dirs = sorted(
        [p for p in images_root.iterdir() if p.is_dir() and p.name.startswith("Pat")],
        key=patient_sort_key
    )

    for patient_dir in patient_dirs:
        subdirs = sorted([p for p in patient_dir.iterdir() if p.is_dir()], key=lambda p: p.name.lower())
        found_phase_names = [p.name for p in subdirs]
        expected_present = [p for p in found_phase_names if p in EXPECTED_PHASES]
        unexpected = [p for p in found_phase_names if p not in EXPECTED_PHASES]

        phase_counts: Dict[str, int] = {}
        phase_present: Dict[str, bool] = {}

        total_files = 0
        for phase in EXPECTED_PHASES:
            phase_folder = patient_dir / phase
            present = phase_folder.is_dir()
            count = count_nii_gz_files(phase_folder)

            phase_present[phase] = present
            phase_counts[phase] = count
            total_files += count

            long_rows.append([
                patient_dir.name,
                phase,
                "yes" if present else "no",
                count,
                str(phase_folder),
            ])

        for folder_name in unexpected:
            folder_path = patient_dir / folder_name
            unexpected_rows.append([
                patient_dir.name,
                folder_name,
                count_nii_gz_files(folder_path),
                str(folder_path),
            ])

        overview_row = [
            patient_dir.name,
            str(patient_dir),
            total_files,
            ", ".join(expected_present) if expected_present else "",
            ", ".join(unexpected) if unexpected else "",
        ]

        for phase in EXPECTED_PHASES:
            overview_row.extend([
                "yes" if phase_present[phase] else "no",
                phase_counts[phase],
            ])

        overview_rows.append(overview_row)

    return overview_rows, long_rows, unexpected_rows


def style_header(ws, row: int = 1):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, extra: int = 2, max_width: int = 40):
    for col_cells in ws.columns:
        length = 0
        column = col_cells[0].column
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[get_column_letter(column)].width = min(length + extra, max_width)


def write_sheet(ws, headers: List[str], rows: List[List]):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws)
    ws.freeze_panes = "A2"
    auto_width(ws)


def build_workbook(overview_rows, long_rows, unexpected_rows) -> Workbook:
    wb = Workbook()

    ws_overview = wb.active
    ws_overview.title = "overview"
    overview_headers = [
        "patient",
        "patient_path",
        "total_nii_gz_files",
        "detected_expected_phases",
        "unexpected_phase_folders",
        "arteriell_present", "arteriell_file_count",
        "nativ_present", "nativ_file_count",
        "spaet_present", "spaet_file_count",
        "venous_present", "venous_file_count",
        "non_def_present", "non_def_file_count",
    ]
    write_sheet(ws_overview, overview_headers, overview_rows)

    ws_long = wb.create_sheet("long_format")
    long_headers = [
        "patient",
        "phase",
        "phase_folder_exists",
        "nii_gz_file_count",
        "phase_folder_path",
    ]
    write_sheet(ws_long, long_headers, long_rows)

    ws_unexpected = wb.create_sheet("unexpected_folders")
    unexpected_headers = [
        "patient",
        "unexpected_folder_name",
        "nii_gz_file_count",
        "folder_path",
    ]
    write_sheet(ws_unexpected, unexpected_headers, unexpected_rows)

    return wb


def main():
    overview_rows, long_rows, unexpected_rows = collect_summary(IMAGES_ROOT)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(overview_rows, long_rows, unexpected_rows)
    wb.save(OUTPUT_XLSX)

    print(f"Summary written to: {OUTPUT_XLSX}")
    print(f"Patients found: {len(overview_rows)}")
    print(f"Phase rows written: {len(long_rows)}")
    print(f"Unexpected folders found: {len(unexpected_rows)}")

if __name__ == "__main__":
    main()
